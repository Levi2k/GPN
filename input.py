import json
from openpyxl import Workbook, load_workbook

def making_import_data():
	PATH_TO_BASE = r'C:\Users\Shakin.VYu\Desktop\База данных ВНР - чистовик.xlsx'
	PATH_INPUT_VNR = r'C:\Users\Shakin.VYu\Desktop\code\Приложение ВНР\VNR.json'
	print('Загрузка базы данных')
	wb_base = load_workbook(PATH_TO_BASE, data_only=True)
	ws_base = wb_base['База данных']
	print('Загрузка завершена. Начинаю формирование словаря...')
	data = {}
	for rows_base in range (3, ws_base.max_row + 1):
		field = ws_base.cell(rows_base, 4).value
		if field not in data:
			data[field] = {}
		if field is None:
			continue
		well = str(ws_base.cell(rows_base, 2).value)
		if well not in data[field]:
			data[field][well] = {}
		data[field][well]['Дата запуска на ВНР'] = str(ws_base.cell(rows_base, 23).value)
	with open(PATH_INPUT_VNR, 'w', encoding='windows-1251') as file:
		json.dump(data, file, indent=4, ensure_ascii=False)
	print('Готово')

if __name__ == '__main__':
	making_import_data()