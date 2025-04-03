import pandas as pd
from openpyxl import Workbook, load_workbook
from pprint36 import pprint
class VNR_reader:

	def __init__(self, path_VNR_GPN, path_VNR_Base):
		self.path_VNR_GPN = path_VNR_GPN
		self.path_base = path_VNR_Base
		self.data = {}
	
	def load_vnr(self):
		print('Подгружаю данные')
		self.wb_vnr = load_workbook(self.path_VNR_GPN, data_only = True)
		self.sheets = self.wb_vnr.sheetnames
		print(self.sheets[-1])
		print('Данные загружены, приступаю к обработке')
		
	def making_database(self):
		sheets =[]
		for sheet in self.wb_vnr.worksheets:
			sheets.append(sheet)
		for i in range(-1, -6):
			self.ws_vnr  = self.wb_vnr[f'{sheets[i].title}']
			print(f'Работаем с датой {sheet}')
			for row_vnr in range(5, self.ws_vnr.max_row):
				if str(self.ws_vnr.cell(row_vnr, 2).value).find('скв') != -1:
					print(1)
					break
				if isinstance(self.ws_vnr.cell(row_vnr, 5).value, str) == False:
						continue
				self.date = self.ws_vnr.title
				print(self.date)				
				if self.date not in self.data:
					self.data[self.date] = {}
				self.field = str(self.ws_vnr.cell(row_vnr, 2).value)
				print(self.field)
				if self.field in [None, 'None']:
					i = 1
					while True:
						self.field = str(self.ws_vnr.cell(row_vnr-i, 2).value) 
						i+=1
						if self.field not in ['None']:
							break
				if self.field not in self.data[self.date]:
					self.data[self.date][self.field] = {}
				self.well = self.ws_vnr.cell(row_vnr, 3).value
				if self.well not in self.data[self.date][self.field]:
					self.data[self.date][self.field][self.well] = {}
					self.data[self.date][self.field][self.well]['Куст'] = self.ws_vnr.cell(row_vnr, 4).value
					self.data[self.date][self.field][self.well]['Вид ремонта'] = self.ws_vnr.cell(row_vnr, 5).value
					self.data[self.date][self.field][self.well]['Кол.дней предв. ВНР'] = self.ws_vnr.cell(row_vnr, 6).value
					self.data[self.date][self.field][self.well]['Остановочный дебит жидкости, м3'] = self.ws_vnr.cell(row_vnr, 7).value
					self.data[self.date][self.field][self.well]['Остановочная обводненность, %'] = self.ws_vnr.cell(row_vnr, 8).value
					self.data[self.date][self.field][self.well]['Остановочный дебит нефти, тн.'] = self.ws_vnr.cell(row_vnr, 9).value
					self.data[self.date][self.field][self.well]['Ожидаемый дебит жидкости, м3'] = self.ws_vnr.cell(row_vnr, 10).value
					self.data[self.date][self.field][self.well]['Ожидаемая обводненность, %'] = self.ws_vnr.cell(row_vnr, 11).value
					self.data[self.date][self.field][self.well]['Ожидаемый дебит нефти, тн.'] = self.ws_vnr.cell(row_vnr, 12).value
					self.data[self.date][self.field][self.well]['Прогнозный дебит жидкости, м3'] = self.ws_vnr.cell(row_vnr, 13).value
					self.data[self.date][self.field][self.well]['Прогнозная обводненность, %'] = self.ws_vnr.cell(row_vnr, 14).value
					self.data[self.date][self.field][self.well]['Прогнозный дебит нефти, тн.'] = self.ws_vnr.cell(row_vnr, 15).value
					self.data[self.date][self.field][self.well]['Текущий дебит жидкости, м3'] = self.ws_vnr.cell(row_vnr, 16).value
					self.data[self.date][self.field][self.well]['Текущая обводненность, %'] = self.ws_vnr.cell(row_vnr, 17).value
					self.data[self.date][self.field][self.well]['Текущий дебит нефти, тн.'] = self.ws_vnr.cell(row_vnr, 18).value
					self.data[self.date][self.field][self.well]['Разница с предыдущими сутками по дебиту жидкости, м3'] = self.ws_vnr.cell(row_vnr, 19).value
					self.data[self.date][self.field][self.well]['Разница с предыдущими сутками по дебиту нефти, тн.'] = self.ws_vnr.cell(row_vnr, 20).value
					self.data[self.date][self.field][self.well]['Недостижение по дебиту жидкости, м3'] = self.ws_vnr.cell(row_vnr, 21).value
					self.data[self.date][self.field][self.well]['Недостижение по дебиту нефти, тн.'] = self.ws_vnr.cell(row_vnr, 22).value
					self.data[self.date][self.field][self.well]['Дата запуска на ВНР'] = self.ws_vnr.cell(row_vnr, 23).value
					self.data[self.date][self.field][self.well]['Дни на ВНР'] = self.ws_vnr.cell(row_vnr, 24).value
					self.data[self.date][self.field][self.well]['Дата запуска по фонду, график'] = self.ws_vnr.cell(row_vnr, 25).value
					self.data[self.date][self.field][self.well]['Дата запуска по фонду, прогноз'] = self.ws_vnr.cell(row_vnr, 26).value
					self.data[self.date][self.field][self.well]['Давление на приеме, атм.'] = self.ws_vnr.cell(row_vnr, 27).value
					self.data[self.date][self.field][self.well]['Отбор, м3'] = self.ws_vnr.cell(row_vnr, 28).value
					try:
							self.data[self.date][self.field][self.well]['Примечание'] = 'Закачка ' + re.findall('\d+м3', str(self.ws_vnr.cell(row_vnr, 29).value))[0]
							
					except Exception as e:
							self.data[self.date][self.field][self.well]['Примечание'] = self.ws_vnr.cell(row_vnr, 29).value
						
		pprint(self.data)
		print(len(self.data))
		self.df = pd.DataFrame.from_dict(self.data, orient = 'index')
		self.df.to_excel(r'C:\Users\Shakin.VYu\Desktop\1.xlsx', engine = 'openpyxl')
		
			
			
		
	def Load_database(self):
		dates = []
		self.wb_vnr_base = load_workbook(self.path_base)
		self.ws_vnr_base  = self.wb_vnr_base['База данных']
		start_row = self.ws_vnr_base.max_row 
		for date, fields_dict in self.data.items():
			cases=[]
			for field, wells_dict in fields_dict.items():
				for well, params_dict in wells_dict.items():
					pprint(well)
					pprint(params_dict)
					self.ws_vnr_base.cell(start_row, 1).value = date
					self.ws_vnr_base.cell(start_row, 2).value = well
					self.ws_vnr_base.cell(start_row, 3).value = params_dict['Куст']
					self.ws_vnr_base.cell(start_row, 4).value = field
					self.ws_vnr_base.cell(start_row, 5).value = params_dict['Вид ремонта']
					self.ws_vnr_base.cell(start_row, 6).value = params_dict['Кол.дней предв. ВНР']
					self.ws_vnr_base.cell(start_row, 7).value = params_dict['Остановочный дебит жидкости, м3']
					self.ws_vnr_base.cell(start_row, 8).value = params_dict['Остановочная обводненность, %']
					self.ws_vnr_base.cell(start_row, 9).value = params_dict['Остановочный дебит нефти, тн.']
					self.ws_vnr_base.cell(start_row, 10).value = params_dict['Ожидаемый дебит жидкости, м3']
					self.ws_vnr_base.cell(start_row, 11).value = params_dict['Ожидаемая обводненность, %']
					self.ws_vnr_base.cell(start_row, 12).value = params_dict['Ожидаемый дебит нефти, тн.']
					self.ws_vnr_base.cell(start_row, 13).value = params_dict['Прогнозный дебит жидкости, м3']
					self.ws_vnr_base.cell(start_row, 14).value = params_dict['Прогнозная обводненность, %']
					self.ws_vnr_base.cell(start_row, 15).value = params_dict['Прогнозный дебит нефти, тн.']
					self.ws_vnr_base.cell(start_row, 16).value = params_dict['Текущий дебит жидкости, м3']
					self.ws_vnr_base.cell(start_row, 17).value = params_dict['Текущая обводненность, %']
					self.ws_vnr_base.cell(start_row, 18).value = params_dict['Текущий дебит нефти, тн.']
					self.ws_vnr_base.cell(start_row, 19).value = params_dict['Разница с предыдущими сутками по дебиту жидкости, м3']
					self.ws_vnr_base.cell(start_row, 20).value = params_dict['Разница с предыдущими сутками по дебиту нефти, тн.']
					self.ws_vnr_base.cell(start_row, 21).value = params_dict['Недостижение по дебиту жидкости, м3']
					self.ws_vnr_base.cell(start_row, 22).value = params_dict['Недостижение по дебиту нефти, тн.']
					self.ws_vnr_base.cell(start_row, 23).value = params_dict['Дата запуска на ВНР']
					self.ws_vnr_base.cell(start_row, 24).value = params_dict['Дни на ВНР']
					self.ws_vnr_base.cell(start_row, 25).value = params_dict['Дата запуска по фонду, график']
					self.ws_vnr_base.cell(start_row, 26).value = params_dict['Дата запуска по фонду, прогноз']
					self.ws_vnr_base.cell(start_row, 27).value = params_dict['Давление на приеме, атм.' ]
					self.ws_vnr_base.cell(start_row, 28).value = params_dict['Отбор, м3']
					self.ws_vnr_base.cell(start_row, 29).value = params_dict['Примечание']
					
					start_row+=1
						
		self.wb_vnr_base.save(self.path_base)
			
			
			
	def main(self):
		self.load_vnr()
		self.making_database()
		self.Load_database()
		
		
if __name__ == '__main__':
	# start_date = str(input('Введите дату начала формирования базы данных: DD.MM.YY\n'))
	print('Начинаем работу')
	path_VNR_GPN = r'C:\Users\Shakin.VYu\Desktop\Сводки\ВНР ГПН-Восток.xlsx'
	# path_VNR_GPN = r'C:\Users\Shakin.VYu\Desktop\ВНР.xlsx'
	# path_VNR_Base = r'C:\Users\Shakin.VYu\Desktop\База данных ВНР.xlsx'
	path_VNR_Base = r'C:\Users\Shakin.VYu\Desktop\База данных ВНР.xlsx'
	VNR_reader = VNR_reader(path_VNR_GPN, path_VNR_Base)
	VNR_reader.main()