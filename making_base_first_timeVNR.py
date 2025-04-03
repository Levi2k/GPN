import pandas as pd
from openpyxl import Workbook, load_workbook
from pprint36 import pprint
import re
from datetime import datetime
class VNR_reader:
	
	def __init__(self, path_VNR_COMPANY, path_VNR_Base):
		self.path_VNR_COMPANY = path_VNR_COMPANY
		self.path_base = path_VNR_Base
		self.data = {}
		self.exceptions = ['32.01', '29.02', '32.03', '31.04', '32.05', '31,06', '32.07', '32.08', '31.09', '32.10', '31.11', '32.12']
		
	def load_vnr(self):
		print('Подгружаю данные')
		self.wb_vnr = load_workbook(self.path_VNR_COMPANY, data_only = True)
		self.wb_base = load_workbook(self.path_base, data_only = True)
		print('Данные загружены, приступаю к обработке')
		
	def Parsing_data(self, sheet):
		self.ws_vnr  = self.wb_vnr[f'{sheet.title}']
		print(f'Работаем с датой {sheet}')
		for row_vnr in range(5, self.ws_vnr.max_row):
			if str(self.ws_vnr.cell(row_vnr, 2).value).find('скв') != -1:
				# print(1)
				break
			if isinstance(self.ws_vnr.cell(row_vnr, 5).value, str) == False:
					continue
			self.date = self.ws_vnr.title
			print(self.date)				
			if self.date not in self.data:
				self.data[self.date] = {}
			self.field = str(self.ws_vnr.cell(row_vnr, 2).value)
			# print(self.field)
			if self.field in [None, 'None']:
				i = 1
				while True:
					self.field = str(self.ws_vnr.cell(row_vnr-i, 2).value) 
					i+=1
					if self.field not in ['None']:
						break
			if self.field not in self.data[self.date]:
				self.data[self.date][self.field] = {}
			self.well = self.ws_vnr.cell(row_vnr, 3).value
			if self.well not in self.data[self.date][self.field]:
				self.data[self.date][self.field][self.well] = {}
				self.data[self.date][self.field][self.well]['Куст'] = self.ws_vnr.cell(row_vnr, 4).value
				self.data[self.date][self.field][self.well]['Вид ремонта'] = self.ws_vnr.cell(row_vnr, 5).value
				self.data[self.date][self.field][self.well]['Кол.дней предв. ВНР'] = self.ws_vnr.cell(row_vnr, 6).value
				self.data[self.date][self.field][self.well]['Остановочный дебит жидкости, м3'] = self.ws_vnr.cell(row_vnr, 7).value
				self.data[self.date][self.field][self.well]['Остановочная обводненность, %'] = self.ws_vnr.cell(row_vnr, 8).value
				self.data[self.date][self.field][self.well]['Остановочный дебит нефти, тн.'] = self.ws_vnr.cell(row_vnr, 9).value
				self.data[self.date][self.field][self.well]['Ожидаемый дебит жидкости, м3'] = self.ws_vnr.cell(row_vnr, 10).value
				self.data[self.date][self.field][self.well]['Ожидаемая обводненность, %'] = self.ws_vnr.cell(row_vnr, 11).value
				self.data[self.date][self.field][self.well]['Ожидаемый дебит нефти, тн.'] = self.ws_vnr.cell(row_vnr, 12).value
				self.data[self.date][self.field][self.well]['Прогнозный дебит жидкости, м3'] = self.ws_vnr.cell(row_vnr, 13).value
				self.data[self.date][self.field][self.well]['Прогнозная обводненность, %'] = self.ws_vnr.cell(row_vnr, 14).value
				self.data[self.date][self.field][self.well]['Прогнозный дебит нефти, тн.'] = self.ws_vnr.cell(row_vnr, 15).value
				self.data[self.date][self.field][self.well]['Текущий дебит жидкости, м3'] = self.ws_vnr.cell(row_vnr, 16).value
				self.data[self.date][self.field][self.well]['Текущая обводненность, %'] = self.ws_vnr.cell(row_vnr, 17).value
				self.data[self.date][self.field][self.well]['Текущий дебит нефти, тн.'] = self.ws_vnr.cell(row_vnr, 18).value
				self.data[self.date][self.field][self.well]['Разница с предыдущими сутками по дебиту жидкости, м3'] = self.ws_vnr.cell(row_vnr, 19).value
				self.data[self.date][self.field][self.well]['Разница с предыдущими сутками по дебиту нефти, тн.'] = self.ws_vnr.cell(row_vnr, 20).value
				self.data[self.date][self.field][self.well]['Недостижение по дебиту жидкости, м3'] = self.ws_vnr.cell(row_vnr, 21).value
				self.data[self.date][self.field][self.well]['Недостижение по дебиту нефти, тн.'] = self.ws_vnr.cell(row_vnr, 22).value
				self.data[self.date][self.field][self.well]['Дата запуска на ВНР'] = self.ws_vnr.cell(row_vnr, 23).value
				self.data[self.date][self.field][self.well]['Дни на ВНР'] = self.ws_vnr.cell(row_vnr, 24).value
				self.data[self.date][self.field][self.well]['Дата запуска по фонду, график'] = self.ws_vnr.cell(row_vnr, 25).value
				self.data[self.date][self.field][self.well]['Дата запуска по фонду, прогноз'] = self.ws_vnr.cell(row_vnr, 26).value
				self.data[self.date][self.field][self.well]['Давление на приеме, атм.'] = self.ws_vnr.cell(row_vnr, 28).value
				self.data[self.date][self.field][self.well]['Отбор, м3'] = self.ws_vnr.cell(row_vnr, 29).value
				color = self.ws_vnr.cell(row_vnr, 29).fill.start_color.index
				if color == 'FF92D050':
					self.data[self.date][self.field][self.well]['Метка запуска по фонду'] = 'Запуск по фонду'
				else:
					self.data[self.date][self.field][self.well]['Метка запуска по фонду'] = 'ВНР'
				try:
					self.data[self.date][self.field][self.well]['Примечание'] = 'Закачка ' + re.findall('\d+м3', str(self.ws_vnr.cell(row_vnr, 29).value))[0]
				except Exception as e:
					self.data[self.date][self.field][self.well]['Примечание'] = self.ws_vnr.cell(row_vnr, 30).value
						
			
		# pprint(self.data)
		# print(len(self.data))
		return self.data
			
		
	def Load_database(self, data):
		dates = []
		self.ws_vnr_base  = self.wb_base['База данных']
		try:
			start_row = self.start_del_row
		except Exception as e:
			start_row = self.ws_vnr_base.max_row
			
		for date, fields_dict in self.data.items():
			cases=[]
			for field, wells_dict in fields_dict.items():
				for well, params_dict in wells_dict.items():
					# pprint(well)
					# pprint(params_dict)
					self.ws_vnr_base.cell(start_row, 1).value = date
					self.ws_vnr_base.cell(start_row, 2).value = well
					self.ws_vnr_base.cell(start_row, 3).value = params_dict['Куст']
					self.ws_vnr_base.cell(start_row, 4).value = field
					self.ws_vnr_base.cell(start_row, 5).value = params_dict['Вид ремонта']
					self.ws_vnr_base.cell(start_row, 6).value = params_dict['Кол.дней предв. ВНР']
					self.ws_vnr_base.cell(start_row, 7).value = params_dict['Остановочный дебит жидкости, м3']
					self.ws_vnr_base.cell(start_row, 8).value = params_dict['Остановочная обводненность, %']
					self.ws_vnr_base.cell(start_row, 9).value = params_dict['Остановочный дебит нефти, тн.']
					self.ws_vnr_base.cell(start_row, 10).value = params_dict['Ожидаемый дебит жидкости, м3']
					self.ws_vnr_base.cell(start_row, 11).value = params_dict['Ожидаемая обводненность, %']
					self.ws_vnr_base.cell(start_row, 12).value = params_dict['Ожидаемый дебит нефти, тн.']
					self.ws_vnr_base.cell(start_row, 13).value = params_dict['Прогнозный дебит жидкости, м3']
					self.ws_vnr_base.cell(start_row, 14).value = params_dict['Прогнозная обводненность, %']
					self.ws_vnr_base.cell(start_row, 15).value = params_dict['Прогнозный дебит нефти, тн.']
					self.ws_vnr_base.cell(start_row, 16).value = params_dict['Текущий дебит жидкости, м3']
					self.ws_vnr_base.cell(start_row, 17).value = params_dict['Текущая обводненность, %']
					self.ws_vnr_base.cell(start_row, 18).value = params_dict['Текущий дебит нефти, тн.']
					self.ws_vnr_base.cell(start_row, 19).value = params_dict['Разница с предыдущими сутками по дебиту жидкости, м3']
					self.ws_vnr_base.cell(start_row, 20).value = params_dict['Разница с предыдущими сутками по дебиту нефти, тн.']
					self.ws_vnr_base.cell(start_row, 21).value = params_dict['Недостижение по дебиту жидкости, м3']
					self.ws_vnr_base.cell(start_row, 22).value = params_dict['Недостижение по дебиту нефти, тн.']
					self.ws_vnr_base.cell(start_row, 23).value = params_dict['Дата запуска на ВНР']
					self.ws_vnr_base.cell(start_row, 24).value = params_dict['Дни на ВНР']
					self.ws_vnr_base.cell(start_row, 25).value = params_dict['Дата запуска по фонду, график']
					self.ws_vnr_base.cell(start_row, 26).value = params_dict['Дата запуска по фонду, прогноз']
					self.ws_vnr_base.cell(start_row, 27).value = params_dict['Давление на приеме, атм.' ]
					self.ws_vnr_base.cell(start_row, 28).value = params_dict['Отбор, м3']
					self.ws_vnr_base.cell(start_row, 29).value = params_dict['Примечание']
					self.ws_vnr_base.cell(start_row, 30).value = params_dict['Метка запуска по фонду']
					
					start_row += 1
						
		self.wb_base.save(self.path_base)
			
	def Clearing_Database_from_start_date(self, start_date):
			
			self.start_date = datetime.strptime(start_date, "%d.%m.%y")
			self.ws_vnr =  self.wb_base['База данных']
			max_row = self.ws_vnr.max_row
			print('Начинаю очистку данных\n')
			try:
				for rows_base in range(3, self.ws_vnr.max_row):
					print(rows_base)
					date_str = self.ws_vnr.cell(rows_base, 1).value
					if date_str == '29.02.25':
						continue
					date = datetime.strptime(self.ws_vnr.cell(rows_base, 1).value, "%d.%m.%y")
					if date >= self.start_date:
						self.start_del_row = rows_base
						print(f'Очистка данных начнется с {date}, находящейся в строке № {self.start_del_row}')
						break
				self.del_amount = self.ws_vnr.max_row - self.start_del_row + 1
				self.ws_vnr.delete_rows(idx = self.start_del_row, amount = self.del_amount)
				self.wb_base.save(self.path_base)
			except Exception as e:
				print(e)
				print(f'Дат после {datetime.strftime(self.start_date, "%Y-%m-%d")} в сводном ВСП не обнаружено')
					
	def Clearing_full_base(self):
		self.ws_vnr =  self.wb_base['База данных']
		self.ws_vnr.delete_rows(idx = 3, amount = self.ws_vnr.max_row - 2)
			
	def main_non_full_parsing(self, start_date):
		self.load_vnr()
		self.Clearing_Database_from_start_date(start_date)
		self.count = 0
		for sheet in self.wb_vnr.worksheets:
			if any(exception in str(sheet) for exception in self.exceptions):
				continue
			date_in_os = datetime.strptime(sheet.title.rstrip(), "%d.%m.%y")
			if date_in_os >= self.start_date:
				data = self.Parsing_data(sheet)
				self.Load_database(data)
				self.count += 1
			
		if self.count == 0:
			print(f'Данных, начиная с {datetime.strftime(start_date, "%Y-%m-%d")} в ВНР -COMPANY нет')
		
			
		
		
			
	def main_full_parsing(self):
		self.load_vnr()
		for sheet in self.wb_base.worksheets:
			if any(exception in str(sheet) for exception in self.exceptions):
				continue
			data = self.Parsing_data(sheet)
			self.Load_database(data)
		
		
if __name__ == '__main__':
	# start_date = str(input('Введите дату начала формирования базы данных: DD.MM.YY\n'))
	print('Начинаем работу')
	path_VNR_COMPANY = r'C:\Users\Shakin.VYu\Desktop\Сводки\ВНР COMPANY.xlsx' #Файл-ВНР чистовик
	# path_VNR_COMPANY = r'C:\Users\Shakin.VYu\Desktop\ВНР.xlsx' #Файл ВНР для апробации (2 листа)
	path_VNR_Base = r'C:\Users\Shakin.VYu\Desktop\База данных ВНР - чистовик.xlsx' #Файл - База чистовая
	# path_VNR_Base = r'C:\Users\Shakin.VYu\Desktop\1.xlsx' #Файл - База для апробации
	regime = int(input('Будем парсить с самого начала, или с конкретной даты? 1 - С самого начала (все листы файла ВНР -COMPANY), 2 - С конкретной даты\n'))
	if regime == 1:
		VNR_reader = VNR_reader(path_VNR_COMPANY, path_VNR_Base)
		VNR_reader.main_full_parsing()
	elif regime == 2:
		start_date = input('Введите стартовую дату в формате: DD.MM.YY\n')
		VNR_reader = VNR_reader(path_VNR_COMPANY, path_VNR_Base)
		VNR_reader.main_non_full_parsing(start_date)